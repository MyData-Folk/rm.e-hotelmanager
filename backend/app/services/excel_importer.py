import io
import re
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlmodel import Session

from app.models.models import AvailabilityCell, ImportedRate, ImportMetadata
from app.services.availability_exporter import normalize_availability_value


DATE_SCAN_LIMIT = 80
PLAN_CODE_RE = re.compile(r"\b[A-Z0-9]{2,}(?:-[A-Z0-9]+)+\b")

AVAILABILITY_SHEET_KEYWORDS = (
    "availability",
    "avail",
    "dispo",
    "disponibilite",
    "disponibilité",
    "planning",
    "stock",
    "inventory",
)

RATE_SHEET_KEYWORDS = (
    "rate",
    "rates",
    "tarif",
    "tarifs",
    "prix",
    "price",
    "prices",
)

GENERIC_LABELS = {
    "date",
    "dates",
    "jour",
    "jours",
    "room",
    "rooms",
    "chambre",
    "chambres",
    "plan",
    "plans",
    "tarif",
    "tarifs",
    "prix",
    "rate",
    "rates",
}


@dataclass
class ParsedExcel:
    import_id: str
    imported_rates: list[ImportedRate] = field(default_factory=list)
    availability_cells: list[AvailabilityCell] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    sheets: list[dict[str, Any]] = field(default_factory=list)


def parse_excel_date(value: Any) -> str | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, (int, float)) and value > 20000:
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date().isoformat()
        except (OverflowError, ValueError):
            return None

    raw = str(value).strip()
    if not raw:
        return None

    for fmt in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%d.%m.%Y",
        "%d.%m.%y",
        "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            pass

    return None


def parse_price(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        return float(value)

    raw = str(value).strip()
    if not raw:
        return None

    raw = raw.replace("€", "").replace(" ", "").replace(",", ".")

    try:
        return float(raw)
    except ValueError:
        return None


def text_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_for_keyword(value: str) -> str:
    return (
        value.lower()
        .replace("é", "e")
        .replace("è", "e")
        .replace("ê", "e")
        .replace("à", "a")
        .replace("ù", "u")
    )


def classify_sheet(title: str) -> str:
    normalized = normalize_for_keyword(title)

    if any(keyword in normalized for keyword in AVAILABILITY_SHEET_KEYWORDS):
        return "availability"

    if any(keyword in normalized for keyword in RATE_SHEET_KEYWORDS):
        return "rates"

    return "mixed"


def extract_plan_code(values: list[Any], fallback: str | None = None) -> str | None:
    candidates = [text_value(value).upper() for value in values]
    if fallback:
        candidates.append(fallback.upper())

    for candidate in candidates:
        match = PLAN_CODE_RE.search(candidate)
        if match:
            return match.group(0)

    return None


def extract_prefixed_plan_code(value: Any) -> str | None:
    label = text_value(value)
    if not label:
        return None

    prefix = label.split(" - ", 1)[0].strip()
    if PLAN_CODE_RE.fullmatch(prefix.upper()):
        return prefix.upper()

    return extract_plan_code([label])


def extract_room_name(values: list[Any], plan_code: str | None = None) -> str | None:
    labels = []

    for value in values:
        label = text_value(value)
        if not label:
            continue

        normalized = normalize_for_keyword(label)
        if normalized in GENERIC_LABELS:
            continue

        if parse_excel_date(label):
            continue

        if plan_code and plan_code in label.upper():
            label = label.upper().replace(plan_code, "").strip(" -_/")

        if not label or PLAN_CODE_RE.fullmatch(label.upper()):
            continue

        labels.append(label)

    if not labels:
        return None

    return max(labels, key=len)


def useful_value(value: Any) -> bool:
    return text_value(value) != ""


def find_date_header_rows(sheet: Worksheet) -> list[dict[str, Any]]:
    headers = []
    max_row = min(sheet.max_row, DATE_SCAN_LIMIT)

    for row_index in range(1, max_row + 1):
        dates = []
        for col_index in range(1, sheet.max_column + 1):
            parsed_date = parse_excel_date(sheet.cell(row_index, col_index).value)
            if parsed_date:
                dates.append((col_index, parsed_date))

        if dates:
            headers.append({"row": row_index, "dates": dates})

    return headers


def find_date_header_columns(sheet: Worksheet) -> list[dict[str, Any]]:
    headers = []
    max_col = min(sheet.max_column, DATE_SCAN_LIMIT)

    for col_index in range(1, max_col + 1):
        dates = []
        for row_index in range(1, sheet.max_row + 1):
            parsed_date = parse_excel_date(sheet.cell(row_index, col_index).value)
            if parsed_date:
                dates.append((row_index, parsed_date))

        if dates:
            headers.append({"column": col_index, "dates": dates})

    return headers


def add_availability_cell(
    parsed: ParsedExcel,
    seen: set[tuple[str, str, str]],
    hotel_id: str,
    cell_date: str,
    room_name: str | None,
    value: Any,
) -> bool:
    if not room_name:
        return False

    normalized = normalize_availability_value(value)
    if normalized["status"] == "unknown" and normalized["raw_value"] == "":
        return False

    key = (cell_date, room_name, text_value(value).lower())
    if key in seen:
        return False

    seen.add(key)
    parsed.availability_cells.append(
        AvailabilityCell(
            hotel_id=hotel_id,
            import_id=parsed.import_id,
            date=cell_date,
            room_name=room_name,
            raw_value=normalized["raw_value"],
            available_quantity=normalized["available_quantity"],
            status=normalized["status"],
            label=normalized["label"],
        )
    )
    return True


def add_imported_rate(
    parsed: ParsedExcel,
    seen: set[tuple[str, str, str]],
    hotel_id: str,
    cell_date: str,
    room_name: str | None,
    plan_code: str | None,
    value: Any,
) -> bool:
    price = parse_price(value)
    if price is None or price < 0:
        return False

    if not room_name:
        return False

    plan_code = plan_code or "UNKNOWN"
    key = (cell_date, room_name, plan_code)
    if key in seen:
        return False

    seen.add(key)
    parsed.imported_rates.append(
        ImportedRate(
            hotel_id=hotel_id,
            import_id=parsed.import_id,
            date=cell_date,
            room_name=room_name,
            plan_code=plan_code,
            price=price,
            raw_value=text_value(value),
        )
    )
    return True


def should_parse_availability(sheet_kind: str, value: Any) -> bool:
    if sheet_kind == "availability":
        return useful_value(value)

    normalized = normalize_availability_value(value)
    raw = normalized["raw_value"]

    if raw == "x":
        return True

    return normalized["status"] in {"sold_out", "available"} and parse_price(value) is not None


def should_parse_rate(sheet_kind: str, value: Any, plan_code: str | None) -> bool:
    if parse_price(value) is None:
        return False

    if sheet_kind == "availability":
        return False

    return sheet_kind == "rates" or plan_code is not None


def parse_row_oriented_dates(
    sheet: Worksheet,
    sheet_kind: str,
    header: dict[str, Any],
    parsed: ParsedExcel,
    hotel_id: str,
    rate_keys: set[tuple[str, str, str]],
    availability_keys: set[tuple[str, str, str]],
) -> dict[str, int]:
    header_row = header["row"]
    date_columns = header["dates"]
    first_date_column = min(column for column, _date in date_columns)
    rates_count = 0
    availability_count = 0
    current_room_name: str | None = None

    for row_index in range(header_row + 1, sheet.max_row + 1):
        label_values = [
            sheet.cell(row_index, column_index).value
            for column_index in range(1, first_date_column)
        ]
        plan_code = extract_plan_code(label_values, sheet.title)
        room_name = extract_room_name(label_values, plan_code)

        if room_name:
            current_room_name = room_name
        else:
            room_name = current_room_name

        for column_index, cell_date in date_columns:
            value = sheet.cell(row_index, column_index).value
            if not useful_value(value):
                continue

            if should_parse_availability(sheet_kind, value):
                if add_availability_cell(
                    parsed,
                    availability_keys,
                    hotel_id,
                    cell_date,
                    room_name,
                    value,
                ):
                    availability_count += 1

            if should_parse_rate(sheet_kind, value, plan_code):
                if add_imported_rate(
                    parsed,
                    rate_keys,
                    hotel_id,
                    cell_date,
                    room_name,
                    plan_code,
                    value,
                ):
                    rates_count += 1

    return {"imported_rates_count": rates_count, "availability_cells_count": availability_count}


def parse_column_oriented_dates(
    sheet: Worksheet,
    sheet_kind: str,
    header: dict[str, Any],
    parsed: ParsedExcel,
    hotel_id: str,
    rate_keys: set[tuple[str, str, str]],
    availability_keys: set[tuple[str, str, str]],
) -> dict[str, int]:
    header_column = header["column"]
    date_rows = header["dates"]
    first_date_row = min(row for row, _date in date_rows)
    rates_count = 0
    availability_count = 0
    current_room_name: str | None = None

    for column_index in range(header_column + 1, sheet.max_column + 1):
        label_values = [
            sheet.cell(row_index, column_index).value
            for row_index in range(1, first_date_row)
        ]
        plan_code = extract_plan_code(label_values, sheet.title)
        room_name = extract_room_name(label_values, plan_code)

        if room_name:
            current_room_name = room_name
        else:
            room_name = current_room_name

        for row_index, cell_date in date_rows:
            value = sheet.cell(row_index, column_index).value
            if not useful_value(value):
                continue

            if should_parse_availability(sheet_kind, value):
                if add_availability_cell(
                    parsed,
                    availability_keys,
                    hotel_id,
                    cell_date,
                    room_name,
                    value,
                ):
                    availability_count += 1

            if should_parse_rate(sheet_kind, value, plan_code):
                if add_imported_rate(
                    parsed,
                    rate_keys,
                    hotel_id,
                    cell_date,
                    room_name,
                    plan_code,
                    value,
                ):
                    rates_count += 1

    return {"imported_rates_count": rates_count, "availability_cells_count": availability_count}


def is_planning_report(sheet: Worksheet, row_headers: list[dict[str, Any]]) -> bool:
    first_row_headers = [
        header
        for header in row_headers
        if header["row"] == 1 and min(column for column, _date in header["dates"]) >= 4
    ]
    if not first_row_headers:
        return False

    descriptors = {
        normalize_for_keyword(text_value(sheet.cell(row_index, 3).value))
        for row_index in range(2, min(sheet.max_row, 30) + 1)
    }
    return any("left for sale" in value for value in descriptors) and any(
        "price" in value for value in descriptors
    )


def parse_planning_report(
    sheet: Worksheet,
    header: dict[str, Any],
    parsed: ParsedExcel,
    hotel_id: str,
    rate_keys: set[tuple[str, str, str]],
    availability_keys: set[tuple[str, str, str]],
) -> dict[str, int]:
    rates_count = 0
    availability_count = 0
    current_room_name: str | None = None

    for row_index in range(header["row"] + 1, sheet.max_row + 1):
        room_label = text_value(sheet.cell(row_index, 1).value)
        if room_label:
            current_room_name = room_label

        if not current_room_name:
            continue

        descriptor = normalize_for_keyword(text_value(sheet.cell(row_index, 3).value))

        if "left for sale" in descriptor:
            for column_index, cell_date in header["dates"]:
                if add_availability_cell(
                    parsed,
                    availability_keys,
                    hotel_id,
                    cell_date,
                    current_room_name,
                    sheet.cell(row_index, column_index).value,
                ):
                    availability_count += 1
            continue

        if "price" not in descriptor:
            continue

        plan_code = extract_prefixed_plan_code(sheet.cell(row_index, 2).value)
        for column_index, cell_date in header["dates"]:
            if add_imported_rate(
                parsed,
                rate_keys,
                hotel_id,
                cell_date,
                current_room_name,
                plan_code,
                sheet.cell(row_index, column_index).value,
            ):
                rates_count += 1

    return {"imported_rates_count": rates_count, "availability_cells_count": availability_count}


def parse_workbook(hotel_id: str, excel_bytes: bytes) -> ParsedExcel:
    parsed = ParsedExcel(import_id=uuid.uuid4().hex[:12])

    try:
        workbook = load_workbook(io.BytesIO(excel_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"Excel invalide: {exc}") from exc

    rate_keys: set[tuple[str, str, str]] = set()
    availability_keys: set[tuple[str, str, str]] = set()

    for sheet in workbook.worksheets:
        sheet_kind = classify_sheet(sheet.title)
        row_headers = find_date_header_rows(sheet)
        column_headers = find_date_header_columns(sheet)
        sheet_summary = {
            "name": sheet.title,
            "kind": sheet_kind,
            "date_header_rows": len(row_headers),
            "date_header_columns": len(column_headers),
            "imported_rates_count": 0,
            "availability_cells_count": 0,
        }

        if not row_headers and not column_headers:
            parsed.warnings.append(
                f"Feuille '{sheet.title}': aucune série de dates reconnue."
            )
            parsed.sheets.append(sheet_summary)
            continue

        if is_planning_report(sheet, row_headers):
            planning_header = next(header for header in row_headers if header["row"] == 1)
            counts = parse_planning_report(
                sheet,
                planning_header,
                parsed,
                hotel_id,
                rate_keys,
                availability_keys,
            )
            sheet_summary["kind"] = "planning_report"
            sheet_summary["imported_rates_count"] += counts["imported_rates_count"]
            sheet_summary["availability_cells_count"] += counts["availability_cells_count"]
            if (
                sheet_summary["imported_rates_count"] == 0
                and sheet_summary["availability_cells_count"] == 0
            ):
                parsed.warnings.append(
                    f"Feuille '{sheet.title}': rapport Planning detecte mais aucune cellule exploitable."
                )
            parsed.sheets.append(sheet_summary)
            continue

        best_row_date_count = max(
            (len(header["dates"]) for header in row_headers),
            default=0,
        )
        best_column_date_count = max(
            (len(header["dates"]) for header in column_headers),
            default=0,
        )

        if row_headers and best_row_date_count >= best_column_date_count:
            for header in row_headers[:2]:
                counts = parse_row_oriented_dates(
                    sheet,
                    sheet_kind,
                    header,
                    parsed,
                    hotel_id,
                    rate_keys,
                    availability_keys,
                )
                sheet_summary["imported_rates_count"] += counts["imported_rates_count"]
                sheet_summary["availability_cells_count"] += counts[
                    "availability_cells_count"
                ]
        elif column_headers:
            for header in column_headers[:2]:
                counts = parse_column_oriented_dates(
                    sheet,
                    sheet_kind,
                    header,
                    parsed,
                    hotel_id,
                    rate_keys,
                    availability_keys,
                )
                sheet_summary["imported_rates_count"] += counts["imported_rates_count"]
                sheet_summary["availability_cells_count"] += counts[
                    "availability_cells_count"
                ]

        if (
            sheet_summary["imported_rates_count"] == 0
            and sheet_summary["availability_cells_count"] == 0
        ):
            parsed.warnings.append(
                f"Feuille '{sheet.title}': dates détectées mais aucune cellule exploitable."
            )

        parsed.sheets.append(sheet_summary)

    return parsed


def import_excel_file(
    session: Session,
    hotel_id: str,
    filename: str,
    excel_bytes: bytes,
) -> dict[str, Any]:
    parsed = parse_workbook(hotel_id, excel_bytes)

    for imported_rate in parsed.imported_rates:
        session.add(imported_rate)

    for availability_cell in parsed.availability_cells:
        session.add(availability_cell)

    metadata = ImportMetadata(
        hotel_id=hotel_id,
        import_id=parsed.import_id,
        import_type="excel",
        filename=filename,
        rows_count=len(parsed.imported_rates) + len(parsed.availability_cells),
        metadata_json={
            "sheets": parsed.sheets,
            "warnings": parsed.warnings,
            "imported_rates_count": len(parsed.imported_rates),
            "availability_cells_count": len(parsed.availability_cells),
        },
    )
    session.add(metadata)
    session.commit()
    session.refresh(metadata)

    return {
        "message": "Excel importé",
        "hotel_id": hotel_id,
        "import_id": parsed.import_id,
        "filename": filename,
        "imported_rates_count": len(parsed.imported_rates),
        "availability_cells_count": len(parsed.availability_cells),
        "warnings": parsed.warnings,
        "sheets": parsed.sheets,
    }
