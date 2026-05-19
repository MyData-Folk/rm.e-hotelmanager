import io

from openpyxl import Workbook
from sqlmodel import Session, SQLModel, create_engine, select

from app.models.models import AvailabilityCell, ImportedRate, ImportMetadata
from app.services.excel_importer import import_excel_file, parse_workbook


def workbook_bytes(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_parse_workbook_imports_availability_and_rates_from_date_columns():
    workbook = Workbook()

    availability = workbook.active
    availability.title = "Disponibilites"
    availability.append(["Chambre", "2026-06-01", "2026-06-02", "2026-06-03"])
    availability.append(["Double Classique", "x", 0, 7])

    rates = workbook.create_sheet("Tarifs OTA-RO-FLEX")
    rates.append(["Chambre", "2026-06-01", "2026-06-02"])
    rates.append(["Double Classique", 200, "210,50"])

    result = parse_workbook("folkestone", workbook_bytes(workbook))

    assert len(result.availability_cells) == 3
    assert len(result.imported_rates) == 2
    assert result.availability_cells[0].status == "not_available_for_sale"
    assert result.availability_cells[1].status == "sold_out"
    assert result.availability_cells[2].available_quantity == 7
    assert result.imported_rates[0].plan_code == "OTA-RO-FLEX"
    assert result.imported_rates[1].price == 210.50
    assert not result.warnings


def test_parse_workbook_warns_when_no_dates_are_detected():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Notes"
    sheet.append(["Room", "Value"])
    sheet.append(["Double Classique", "No date here"])

    result = parse_workbook("folkestone", workbook_bytes(workbook))

    assert result.imported_rates == []
    assert result.availability_cells == []
    assert result.warnings == ["Feuille 'Notes': aucune série de dates reconnue."]
    assert result.sheets[0]["name"] == "Notes"


def test_parse_workbook_imports_vertical_date_column_layout():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Planning"
    sheet.append(["Date", "Double Classique", "Twin Classique"])
    sheet.append(["2026-06-01", "x", 2])
    sheet.append(["2026-06-02", 0, 5])

    result = parse_workbook("folkestone", workbook_bytes(workbook))

    assert len(result.availability_cells) == 4
    assert {
        (cell.date, cell.room_name)
        for cell in result.availability_cells
    } == {
        ("2026-06-01", "Double Classique"),
        ("2026-06-02", "Double Classique"),
        ("2026-06-01", "Twin Classique"),
        ("2026-06-02", "Twin Classique"),
    }
    assert {cell.available_quantity for cell in result.availability_cells} == {
        None,
        0,
        2,
        5,
    }


def test_parse_workbook_imports_folkestone_planning_report_layout():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Planning"
    sheet.append([
        "FOLKESTONE OPERA - mercredi 13 mai 2026 16:56:58",
        None,
        None,
        "13/05/26",
        "14/05/26",
    ])
    sheet.append(["Double Classique", None, "Left for sale", 3, 0])
    sheet.append([
        "Double Classique",
        "RACK-RO-FLEX - RACK RO FLEX",
        "Price (EUR)",
        161,
        188,
    ])
    sheet.append([
        "Double Classique",
        "OTA-RO-FLEX - OTA RO FLEX",
        "Price (EUR)",
        "151,50",
        178,
    ])

    result = parse_workbook("folkestone", workbook_bytes(workbook))

    assert result.sheets[0]["kind"] == "planning_report"
    assert len(result.availability_cells) == 2
    assert len(result.imported_rates) == 4
    assert {rate.plan_code for rate in result.imported_rates} == {
        "OTA-RO-FLEX",
        "RACK-RO-FLEX",
    }
    assert {rate.room_name for rate in result.imported_rates} == {"Double Classique"}
    assert result.imported_rates[2].price == 151.50


def test_import_excel_file_persists_metadata_rates_and_availability():
    workbook = Workbook()
    availability = workbook.active
    availability.title = "Planning"
    availability.append(["Chambre", "2026-06-01", "2026-06-02"])
    availability.append(["Twin Classique", "x", 3])

    rates = workbook.create_sheet("Prix CWT-BB-FLEX")
    rates.append(["Chambre", "2026-06-01"])
    rates.append(["Twin Classique", 225])

    engine = create_engine("sqlite:///:memory:")
    SQLModel.metadata.create_all(engine)

    with Session(engine) as session:
        summary = import_excel_file(
            session=session,
            hotel_id="folkestone",
            filename="planning.xlsx",
            excel_bytes=workbook_bytes(workbook),
        )

        metadata = session.exec(select(ImportMetadata)).one()
        imported_rates = session.exec(select(ImportedRate)).all()
        availability_cells = session.exec(select(AvailabilityCell)).all()

    assert summary["message"] == "Excel importé"
    assert summary["import_id"] == metadata.import_id
    assert metadata.rows_count == 3
    assert len(imported_rates) == 1
    assert imported_rates[0].plan_code == "CWT-BB-FLEX"
    assert len(availability_cells) == 2
    assert {cell.status for cell in availability_cells} == {
        "not_available_for_sale",
        "available",
    }
